from functools import total_ordering
from utils import *
import pandas as pd
import json # Ensure json is imported
from openpyxl import Workbook # For creating new Excel files
from openpyxl.utils.dataframe import dataframe_to_rows # For writing dataframes to sheets

def generate_excel_for_count_posts(data, platform):
    # 将数据转换为DataFrame
    df = pd.DataFrame.from_dict(data, orient='index')

    if platform == "wb":
        df = df[['total_posts', 'total_replies', 'total', '范围内帖子', '范围内评论', '范围内总数', 'hotel_related_posts', 'hotel_related_replies', '有效数据数量', '最早时间', '最晚时间', '内容不完整帖子占比有效帖子']]
    elif platform == "xhs":
        df = df[['total_posts', 'total_replies', 'total', '范围内帖子', '范围内评论', '范围内总数', '软文数量', 'hotel_related_posts', 'hotel_related_replies', '有效数据数量', '最早时间', '最晚时间']]
    else:
        df = df[['total_posts', 'total_replies', 'total', '范围内帖子', '范围内评论', '范围内总数', 'hotel_related_posts', 'hotel_related_replies', '有效数据数量', '最早时间', '最晚时间']]

    # 使用ExcelWriter来处理多个sheet的写入
    try:
        with pd.ExcelWriter('analysis_result/数据量统计.xlsx', mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=f'{platform}帖子统计')
    except FileNotFoundError:
        # 如果文件不存在，创建新文件
        df.to_excel('analysis_result/数据量统计.xlsx', sheet_name=f'{platform}帖子统计')

    print("Excel文件已生成：analysis_result/数据量统计.xlsx")

def count_posts(platform):
    filter = PostsFilter()
    raw_data = get_raw_data(f'raw_data/{platform}.json')
    filtered_data = filter.filter_by_time(raw_data)
    
    if not raw_data:
        return None
    
    count = {}
    for hotel in raw_data:
        latest_time = 0
        earliest_time = float('inf')

        posts_count = len(hotel["posts"])
        total_replies = 0
        for post in hotel["posts"]:
            total_replies += len(post["replies"])
        if len(hotel["posts"]) == 0:
            print(f"警告：酒店 '{hotel['hotel']}' 没有帖子，跳过统计。")
            continue
        for post in hotel["posts"]:
            timestamp = datetime.strptime(post["timestamp"], '%Y-%m-%d %H:%M').timestamp()
            if timestamp < earliest_time:
                earliest_time = timestamp
            if timestamp > latest_time:
                latest_time = timestamp
        tmp = {
            "total_posts": posts_count,
            "total_replies": total_replies,
            "total": posts_count + total_replies,
            "最早时间": datetime.fromtimestamp(earliest_time).strftime('%Y-%m-%d %H:%M'),
            "最晚时间": datetime.fromtimestamp(latest_time).strftime('%Y-%m-%d %H:%M'),
        }
        count[hotel["hotel"]] = tmp
        
    # 读取analyzed数据，只需读取一次
    analyzed_data = None
    try:
        with open(f'analysis_result/{platform}_analyzed.json', 'r', encoding='utf-8') as f:
            analyzed_data = json.load(f)
    except FileNotFoundError:
        print(f"警告：无法找到 {platform}_analyzed.json 文件，将无法统计有效数据和软文数量。")
    except Exception as e:
        print(f"读取 {platform}_analyzed.json 文件时出错: {e}")

    # 将 analyzed_data 转换为字典以便快速查找
    analyzed_hotel_map = {h['hotel']: h for h in analyzed_data} if analyzed_data else {}

    for hotel in filtered_data:
        hotel_name = hotel["hotel"]
        replies_count = 0
        hotel_related_posts = 0
        hotel_related_replies = 0
        ad_count = 0 # 初始化 ad_count
        incomplete_posts = 0
        
        posts_count = len(hotel["posts"])
        if hotel_name not in count:
             print(f"警告：酒店 '{hotel_name}' 在 raw_data 中未找到对应条目，跳过统计。")
             continue
             
        count[hotel_name]["范围内帖子"] = posts_count
        
        # 获取对应的 analyzed_hotel 数据
        analyzed_hotel = analyzed_hotel_map.get(hotel_name)

        # --- 修正开始 ---
        # 在这里一次性计算该酒店的软文数量 (ad_count)
        if analyzed_hotel and platform == "xhs":
            for analyzed_post in analyzed_hotel.get("posts", []):
                if analyzed_post.get("is_ad", False):
                    ad_count += 1
            count[hotel_name]["软文数量"] = ad_count
        elif platform == "xhs":
             count[hotel_name]["软文数量"] = 0 # 如果没有 analyzed_hotel，软文数为0
        # --- 修正结束 ---
                    
        for post in hotel["posts"]:
            replies_count += len(post["replies"])
            
            if "...全文" in post["content"]:
                incomplete_posts += 1
            
            # 统计酒店相关帖子和评论 (使用 analyzed_hotel)
            if analyzed_hotel:
                # 查找与当前 post 匹配的 analyzed_post
                matching_analyzed_post = None
                for ap in analyzed_hotel.get("posts", []):
                    if ap.get("link") == post.get("link"): # 假设 link 是唯一标识符
                        matching_analyzed_post = ap
                        break
                
                if matching_analyzed_post and matching_analyzed_post.get("is_hotel_related", False):
                    hotel_related_posts += 1
                    # 统计该帖子下的酒店相关评论
                    for reply in matching_analyzed_post.get("replies", []):
                        if reply.get("is_hotel_related", False):
                            hotel_related_replies += 1
                
        count[hotel_name]["范围内评论"] = replies_count
        count[hotel_name]["范围内总数"] = posts_count + replies_count
        # 有效数据数量统计逻辑不变
        count[hotel_name]["hotel_related_posts"] = hotel_related_posts
        count[hotel_name]["hotel_related_replies"] = hotel_related_replies
        count[hotel_name]["有效数据数量"] = hotel_related_posts + hotel_related_replies - ad_count
        
        # 内容不完整帖子占比统计逻辑不变
        if platform == "wb":
            if posts_count == 0:
                count[hotel_name]["内容不完整帖子占比有效帖子"] = "0%"
            else:
                count[hotel_name]["内容不完整帖子占比有效帖子"] = f"{round(incomplete_posts/posts_count * 100)}%"
        
    generate_excel_for_count_posts(count, platform)
    print("共有{}个酒店".format(len(count)))
    print("酒店列表： ", list(count.keys()))


def merge_xhs_posts(posts_path, comments_path, new_hotel_name):
    with open(posts_path, "r", encoding="utf-8") as f:
        posts = json.load(f)
    with open(comments_path, "r", encoding="utf-8") as f:
        comments = json.load(f)

    # 使用字典去重posts
    unique_posts = {}
    for post in posts:
        if post['note_id'] not in unique_posts:
            unique_posts[post['note_id']] = post

    # 使用字典去重comments
    unique_comments = {}
    for comment in comments:
        if comment['comment_id'] not in unique_comments:
            unique_comments[comment['comment_id']] = comment

    # 将评论按note_id分组
    comments_by_post = {}
    for comment in unique_comments.values():
        note_id = comment['note_id']
        if note_id not in comments_by_post:
            comments_by_post[note_id] = []
        # 格式化评论时间
        comment_time = datetime.fromtimestamp(int(comment.get('create_time', 0)) / 1000).strftime('%Y-%m-%d %H:%M')
        formatted_comment = {
            "commenter_name": comment.get('nickname', ''),
            "comment_content": comment.get('content', ''),
            "commenter_link": comment.get('profile_url', ''),
            "comment_time": comment_time
        }
        comments_by_post[note_id].append(formatted_comment)

    formatted_posts = []
    for post in posts:
        # 格式化帖子时间
        post_time = datetime.fromtimestamp(int(post.get('last_update_time', 0))/1000).strftime('%Y-%m-%d %H:%M')
        formatted_post = {
            "note_id": post['note_id'],
            "content": post.get('title', '') + post.get('desc', ''),
            "timestamp": post_time,
            "link": post.get('note_url', ''),
            "replies": comments_by_post.get(post['note_id'], [])
        }
        formatted_posts.append(formatted_post)
            
    with open("raw_data/xhs.json", "r", encoding="utf-8") as f:
        xhs_data = json.load(f)
    
    # 将现有的 xhs_data 转换为以 hotel 为键的字典
    hotel_dict = {}
    for item in xhs_data:
        hotel_name = item['hotel']
        if hotel_name not in hotel_dict:
            hotel_dict[hotel_name] = item
        else:
            # 如果已存在该酒店，将 posts 添加到现有列表中
            hotel_dict[hotel_name]['posts'].extend(item['posts'])
        
    if new_hotel_name in hotel_dict:
        print(f"酒店{new_hotel_name}已存在，将新帖子添加到现有数据中...")
        print(f"现有帖子数量：{len(hotel_dict[new_hotel_name]['posts'])}")
        existing_note_ids = [post['note_id'] for post in hotel_dict[hotel_name]['posts']]
        for post in formatted_posts:
            if post['note_id'] not in existing_note_ids:
                hotel_dict[new_hotel_name]['posts'].append(post)
        print(f"添加成功！新帖子数量：{len(hotel_dict[new_hotel_name]['posts'])}")
    else:
        # 如果是新酒店，直接添加
        hotel_dict[new_hotel_name] = {
            "hotel": new_hotel_name,
            "posts": formatted_posts
        }
        print(f"新酒店{hotel_name}添加成功！帖子数量：{len(formatted_posts)}")
    
    # 将字典转换回列表形式
    final_data = list(hotel_dict.values())
    
    write_to_json(final_data, "raw_data/xhs.json")
        

def analyze_sentiment_distribution(analyzed_data, keywords_data):
    """
    args:
        analyzed_data: 分析后的json数据 (列表，每个元素是一个酒店的数据字典)
        keywords_data: 关键词数据 (列表，每个元素是一级关键词及其二级关键词的字典)

    统计逻辑：
    遍历所有酒店的帖子及回复，对于帖子提及的关键词，其评论的数量可以在一定程度上反应帖子提及关键词的情感倾向，
    因此在统计帖子提及关键词时，在每个关键词的对应sentiment统计上加上回复数量。
    对于回复本身提及的关键词就只在对应关键词的统计上加1。

    return:
    {
        "酒店名称1": {
            "一级关键词A": {
                "secondary_keywords": {
                    "二级关键词A1": {
                        "sentiment_distribution": { "positive": 0, "negative": 0, "neutral": 0 }
                    },
                    "二级关键词A2": { ... }
                },
                "sentiment_distribution": { "positive": 0, "negative": 0, "neutral": 0 } # 一级关键词A的总情感分布
            },
            "一级关键词B": { ... }
        },
        "酒店名称2": { ... }
    }
    """
    results = {}

    # 创建关键词映射，方便查找二级关键词所属的一级关键词
    primary_keyword_map = {}
    for pk in keywords_data:
        primary_keyword_map[pk['primary_keyword']] = {
            'secondary_keywords': {},
            'sentiment_distribution': {'positive': 0, 'negative': 0, 'neutral': 0}
        }
        for sk in pk['secondary_keywords']:
            primary_keyword_map[pk['primary_keyword']]['secondary_keywords'][sk['keyword']] = {
                'sentiment_distribution': {'positive': 0, 'negative': 0, 'neutral': 0}
            }

    for hotel_entry in analyzed_data:
        hotel_name = hotel_entry['hotel']
        if hotel_name not in results:
            results[hotel_name] = json.loads(json.dumps(primary_keyword_map)) # 深拷贝

        for post in hotel_entry['posts']:
            post_contribution_factor = 1 + len(post['replies'])

            if 'keywords_mentioned' in post and post['keywords_mentioned']:
                # 处理帖子自身的一级关键词
                if 'primary_keyword' in post['keywords_mentioned'] and post['keywords_mentioned']['primary_keyword']:
                    for pk_mention in post['keywords_mentioned']['primary_keyword']:
                        pk_name = pk_mention['keyword']
                        sentiment = pk_mention.get('sentiment', 'neutral').lower()
                        if pk_name in results[hotel_name]:
                            results[hotel_name][pk_name]['sentiment_distribution'][sentiment] += post_contribution_factor

                # 处理帖子自身的二级关键词
                if 'secondary_keyword' in post['keywords_mentioned'] and post['keywords_mentioned']['secondary_keyword']:
                    for sk_mention in post['keywords_mentioned']['secondary_keyword']:
                        sk_name = sk_mention['keyword']
                        sentiment = sk_mention.get('sentiment', 'neutral').lower()
                        for pk_name_iter in results[hotel_name]:
                            if 'secondary_keywords' in results[hotel_name][pk_name_iter] and \
                               sk_name in results[hotel_name][pk_name_iter]['secondary_keywords']:
                                results[hotel_name][pk_name_iter]['secondary_keywords'][sk_name]['sentiment_distribution'][sentiment] += post_contribution_factor
                                # 同时，将二级关键词的情感也计入对应的一级关键词
                                results[hotel_name][pk_name_iter]['sentiment_distribution'][sentiment] += post_contribution_factor
                                break
            
            # 单独处理回复中明确提及的关键词
            for reply in post['replies']:
                if 'keywords_mentioned' in reply and reply['keywords_mentioned']:
                    # 处理回复中的一级关键词
                    if 'primary_keyword' in reply['keywords_mentioned'] and reply['keywords_mentioned']['primary_keyword']:
                        for pk_mention in reply['keywords_mentioned']['primary_keyword']:
                            pk_name = pk_mention['keyword']
                            sentiment = pk_mention.get('sentiment', 'neutral').lower()
                            if pk_name in results[hotel_name]:
                                results[hotel_name][pk_name]['sentiment_distribution'][sentiment] += 1
                    
                    # 处理回复中的二级关键词
                    if 'secondary_keyword' in reply['keywords_mentioned'] and reply['keywords_mentioned']['secondary_keyword']:
                        for sk_mention in reply['keywords_mentioned']['secondary_keyword']:
                            sk_name = sk_mention['keyword']
                            sentiment = sk_mention.get('sentiment', 'neutral').lower()
                            for pk_name_iter in results[hotel_name]:
                                if 'secondary_keywords' in results[hotel_name][pk_name_iter] and \
                                   sk_name in results[hotel_name][pk_name_iter]['secondary_keywords']:
                                    results[hotel_name][pk_name_iter]['secondary_keywords'][sk_name]['sentiment_distribution'][sentiment] += 1
                                    results[hotel_name][pk_name_iter]['sentiment_distribution'][sentiment] += 1
                                    break
    return results

def calculate_scores_for_keyword(sentiment_distribution):
    if sentiment_distribution is None:
        sentiment_distribution = {}
        
    positive = sentiment_distribution.get('positive', 0)
    negative = sentiment_distribution.get('negative', 0)
    neutral = sentiment_distribution.get('neutral', 0)
    
    total_buzz = positive + negative + neutral
    raw_score = positive - negative
    
    sentiment_score_percent = (raw_score / total_buzz) * 100 if total_buzz > 0 else 0
    
    return {'totalBuzz': total_buzz, 'sentimentScorePercent': sentiment_score_percent}
    

def calculate_keyword_scores_and_generate_excel(analyzed_data_list, keywords_data, output_excel_path):
    """
    分析各平台数据中每个酒店的一级和二级关键词的情感数量和得分，并生成Excel文件。

    Args:
        analyzed_data_list (list): 包含各平台分析结果的JSON文件
        keywords_data (list): 关键词配置JSON文件的列表，每个元素是一级关键词及其二级关键词的字典。
        output_excel_path (str): 输出Excel文件的路径。
    """

    all_hotel_data_cumulative = {}
    # 定义列顺序
    column_order = ['一级关键词', '一级正面', '一级中立', '一级负面', '一级关键词情感得分',
                    '二级关键词', '二级正面', '二级中立', '二级负面', '二级关键词情感得分']

    for analyzed_data in analyzed_data_list:
        # 调用已有的情感分布分析函数
        sentiment_distribution_by_hotel = analyze_sentiment_distribution(analyzed_data, keywords_data)

        for hotel_name, hotel_keywords_data in sentiment_distribution_by_hotel.items():
            if hotel_name not in all_hotel_data_cumulative:
                # 初始化酒店条目，深拷贝 keywords_data 结构以包含所有可能的关键词
                all_hotel_data_cumulative[hotel_name] = json.loads(json.dumps(
                    {pk['primary_keyword']: {
                        'sentiment_distribution': {'positive': 0, 'negative': 0, 'neutral': 0},
                        'secondary_keywords': {sk['keyword']: {'sentiment_distribution': {'positive': 0, 'negative': 0, 'neutral': 0}}
                                            for sk in pk['secondary_keywords']}
                    } for pk in keywords_data}
                ))
            
            for pk_name, pk_data_from_current_file in hotel_keywords_data.items():
                if pk_name in all_hotel_data_cumulative[hotel_name]:
                    # 累加一级关键词情感分布
                    current_pk_sentiment = all_hotel_data_cumulative[hotel_name][pk_name]['sentiment_distribution']
                    for sentiment, count in pk_data_from_current_file['sentiment_distribution'].items():
                        current_pk_sentiment[sentiment] = current_pk_sentiment.get(sentiment, 0) + count

                    # 累加二级关键词情感分布
                    if 'secondary_keywords' in pk_data_from_current_file:
                        for sk_name, sk_data_from_current_file in pk_data_from_current_file['secondary_keywords'].items():
                            if sk_name in all_hotel_data_cumulative[hotel_name][pk_name]['secondary_keywords']:
                                current_sk_sentiment = all_hotel_data_cumulative[hotel_name][pk_name]['secondary_keywords'][sk_name]['sentiment_distribution']
                                for sentiment, count in sk_data_from_current_file['sentiment_distribution'].items():
                                    current_sk_sentiment[sentiment] = current_sk_sentiment.get(sentiment, 0) + count

    # 创建Excel工作簿
    wb = Workbook()
    # 删除默认创建的sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for hotel_name, accumulated_hotel_data in all_hotel_data_cumulative.items():
        hotel_specific_rows = [] # 用于存储当前酒店的行数据
        for pk_name, pk_data in accumulated_hotel_data.items():
            pk_scores = calculate_scores_for_keyword(pk_data['sentiment_distribution'])
            pk_sentiment = pk_data['sentiment_distribution']
            
            # 处理一级关键词
            row_data_pk = {
                '一级关键词': pk_name,
                '一级正面': pk_sentiment.get('positive', 0),
                '一级中立': pk_sentiment.get('neutral', 0),
                '一级负面': pk_sentiment.get('negative', 0),
                '一级关键词情感得分': f"{pk_scores['sentimentScorePercent']:.2f}%",
                '二级关键词': '',
                '二级正面': '',
                '二级中立': '',
                '二级负面': '',
                '二级关键词情感得分': ''
            }
            hotel_specific_rows.append(row_data_pk)

            # 处理二级关键词
            if 'secondary_keywords' in pk_data:
                for sk_name, sk_data in pk_data['secondary_keywords'].items():
                    sk_sentiment = sk_data['sentiment_distribution']
                    sk_scores = calculate_scores_for_keyword(sk_sentiment)
                    row_data_sk = {
                        '一级关键词': pk_name, 
                        '一级正面': pk_sentiment.get('positive', 0), 
                        '一级中立': pk_sentiment.get('neutral', 0),
                        '一级负面': pk_sentiment.get('negative', 0),
                        '一级关键词情感得分': f"{pk_scores['sentimentScorePercent']:.2f}%",
                        '二级关键词': sk_name,
                        '二级正面': sk_sentiment.get('positive', 0),
                        '二级中立': sk_sentiment.get('neutral', 0),
                        '二级负面': sk_sentiment.get('negative', 0),
                        '二级关键词情感得分': f"{sk_scores['sentimentScorePercent']:.2f}%",
                    }
                    hotel_specific_rows.append(row_data_sk)
        
        if not hotel_specific_rows: # 如果当前酒店没有数据
            # 创建一个带有表头的空sheet，或者跳过这个酒店的sheet创建
            # 这里选择创建一个带表头的空sheet
            df = pd.DataFrame(columns=column_order) # type: ignore
        else:
            df = pd.DataFrame(hotel_specific_rows)
            # 确保列的顺序与图片一致
            # 检查df是否为空，以及是否包含所有column_order中的列
            if not df.empty and all(col in df.columns for col in column_order):
                df = df[column_order]
            elif df.empty:
                 df = pd.DataFrame(columns=column_order) # type: ignore
            # else: df中缺少某些列，这里可以选择报错或者填充默认值，目前保持原样，如果df非空但列不全，后续写入excel可能会出错或列不对应

        # 创建新的sheet
        # Excel sheet名称有长度限制，且不能包含某些特殊字符，这里简单处理
        safe_hotel_name = "".join(c if c.isalnum() else "_" for c in hotel_name)[:31]
        ws = wb.create_sheet(title=safe_hotel_name)

        # 写入表头
        ws.append(column_order)

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_excel_path)
    print(f"关键词情感得分Excel文件已生成：{output_excel_path}")


if __name__ == "__main__":
    # main()
    # count_posts("xhs")
    # count_posts("wb")
    # count_posts("flyert")

    xhs_analyzed = get_raw_data("analysis_result/xhs_analyzed.json")
    wb_analyzed = get_raw_data("analysis_result/wb_analyzed.json")
    flyert_analyzed = get_raw_data("analysis_result/flyert_analyzed.json")
    analyzed_files = [xhs_analyzed, wb_analyzed, flyert_analyzed]
    keywords_file = get_raw_data("raw_data/keywords.json")
    output_excel = 'analysis_result/keyword_sentiment_scores.xlsx'
    calculate_keyword_scores_and_generate_excel(analyzed_files, keywords_file, output_excel)
